import openpyxl


book = openpyxl.load_workbook("C:\\Users\\brahm\\Desktop\\Python\\Udemy\\rahul_shetty\\project\\seleniumFramework\\TestData\\PythonDemo.xlsx")
sheet = book.active
# cell = sheet.cell(row=1, column=2)
# print(cell.value)

# sheet.cell(row=1, column=2).value = "Tito"
# print(cell.value)
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet['A5'].value)

# Only printing one specific fixed column
# for i in range(1, sheet.max_row+1):
#     print(sheet.cell(row=i, column=1).value)

# printing all rows and columns
# for i in range(1, sheet.max_row+1):
#     for j in range(1, sheet.max_column+1):
#         print(sheet.cell(row=i, column=j).value)

# printing specific column
# for i in range(1, sheet.max_row+1):
#     if sheet.cell(row=i, column=1).value == "Testcase2":
#         for j in range(2, sheet.max_column+1):
#             print(sheet.cell(row=i, column=j).value)


# save the result in dictionary
dict = {}

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase3":
        for j in range(2, sheet.max_column+1):
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(dict)


