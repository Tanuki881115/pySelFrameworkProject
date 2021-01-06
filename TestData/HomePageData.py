import openpyxl


class HomePageData:

    test_homePageData = [{"first_name":"Tito", "last_name":"Novianto", "gender":"Male"},
                         {"first_name":"Rahul", "last_name":"Shetty", "gender":"Male"}
                         ]

    #Using the static method to call a method will not required to put the "self" param in the getTestData method
    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\brahm\\Desktop\\Python\\Udemy\\rahul_shetty\\project\\seleniumFramework\\TestData\\PythonDemo.xlsx")
        sheet = book.active
        dict = {}
        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column+1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[dict]
